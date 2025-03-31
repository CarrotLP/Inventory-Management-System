import tkinter as tk
from tkinter import ttk
from tkinter import font
from tkinter import filedialog
from random import randint
from email.message import EmailMessage
import ssl
import smtplib
import datetime
import sqlite3
import openpyxl
import model

class Login:
    #Instantiate login object
    def __init__(self, master):
        self.frames = []   #Store frames used so they can be called at a later stage
        self.username = ""
        self.loginSuccess = [False]
        self.loginWidgets(master)
        self.__database__ = "./alldata.db"
        
    def clear(self, master):
        #Destroy all widgets in a frame
        for widget in master.winfo_children():
            widget.destroy()

    def dummy(self, frameName, rowNo, columnNo):
        #Dummy label so a blank space can be created
        dummyLbl = ttk.Label(frameName, text=" ")
        dummyLbl.grid(row=rowNo, column=columnNo)

    def etrVal(self, textBox):
        #Get values from entryboxes
        inputVal = textBox.get()
        return inputVal

    def loginWidgets(self, master, attempts=1, error=False):
        #clear screen
        self.clear(master)

        #Top frame
        topFrame = ttk.Frame(master)
        self.frames.append(topFrame)
        topFrame.pack()

        #Error label
        errorLbl = ttk.Label(topFrame, text=f"Login Failed. {6-attempts} attempts left.", foreground="red")
        errorLbl.grid(row=0, column=0)
        errorLbl.grid_remove()
        if error:
            errorLbl.grid(row=0, column=0)

        #Name
        nameLbl = ttk.Label(topFrame, text="太陽油集團（國際）有限公司")
        nameLbl.grid(row=0, column=1, columnspan=3, sticky="W", padx=10)
        
        loginLbl = ttk.Label(topFrame, text="Account login 帳戶登入", anchor="e")
        loginLbl.grid(row=1, column=1, sticky="E", padx=25)

        #Body frame
        bodyFrame = ttk.Frame(master)
        self.frames.append(bodyFrame)
        bodyFrame.pack()

        #Username
        usernameLbl = ttk.Label(bodyFrame, text="Username 帳戶名稱:")
        usernameLbl.grid(row=2, column=0, sticky="W")
        usernameEtr = ttk.Entry(bodyFrame, width="70")
        usernameEtr.grid(row=3, column=0, columnspan=3, sticky="W")

        self.dummy(bodyFrame, 4, 0)

        #Password
        pwLbl = ttk.Label(bodyFrame, text="Password 密碼:")
        pwLbl.grid(row=5, column=0, sticky="W")
        pwEtr = ttk.Entry(bodyFrame, width="70", show="*")
        pwEtr.grid(row=6, column=0, columnspan=3, sticky="W")

        self.dummy(bodyFrame, 7, 0)

        #Enter button
        enterBtn = ttk.Button(bodyFrame, text="Enter 提交", command=lambda : self.checkPw(master, self.etrVal(pwEtr), attempts, username=self.etrVal(usernameEtr)))   
        enterBtn.grid(row=8, column=2, sticky="E")

        #Bind to 'Enter' button on keyboard
        master.bind('<Return>', lambda event: self.checkPw(master, self.etrVal(pwEtr), attempts, username=self.etrVal(usernameEtr), event=True))

    def otp(self, master, topFrame, bodyFrame, emailAdd):
        #Generate 6 random digits
        otpStr = ""
        for j in range (6):
            digit = randint(0,9)
            otpStr += str(digit)
        
        try:
            #Send to user email
            otpSub = "Stock Property Management System OTP"
            otpBody = f"""
            Hi there, if you accidentally received this email, I am sorry it is sent to the wrong recipient!
            Dear staff, 

            The following is your one-time password for logging into the stock property management system. 
            Please do not reply and do not share this code:
            {otpStr}
            """
            self.email(emailAdd, otpSub, otpBody)
            attempt = 0
            self.otpWindow(master, attempt, otpStr)

        except:
            usernameErrLbl = ttk.Label(topFrame, text="Please connect to the Internet", wraplength=170, foreground="red")
            usernameErrLbl.grid(row=0, column=0, sticky="W", padx=5)

            dumm1 = ttk.Label(topFrame, text="                                        ")
            dumm1.grid(row=1, column=0)
            dumm2 = ttk.Label(topFrame, text="                                        ")
            dumm2.grid(row=1, column=2)    

    def otpWindow(self, master, noOfAttempt, otp, error=False):
        #Enter OTP window
        self.clear(master)

        #Create frames
        topFrame = ttk.Frame(master)
        topFrame.pack()
        self.frames.append(topFrame)
        bodyFrame = ttk.Frame(master)
        bodyFrame.pack()
        self.frames.append(bodyFrame)

        #Error label
        if error==True:        
            errorLbl = ttk.Label(topFrame, text=f"OTP incorrect,{5-noOfAttempt} attempts left", wraplength=150, foreground="red")
            errorLbl.grid(row=0, column=0)
            dummy = ttk.Label(topFrame, text="                                      ")
            dummy.grid(row=0, column=3, sticky="E")

        #Company name
        nameLbl = ttk.Label(topFrame, text="太陽油集團（國際）有限公司")
        nameLbl.grid(row=0, column=1, columnspan=2, sticky="W", ipadx=5, ipady=20)

        #Enter OTP
        otpLbl = ttk.Label(bodyFrame, text="OTP:")
        otpLbl.grid(row=2, column=0, sticky="W")
        otpEtr = ttk.Entry(bodyFrame, width="70")
        otpEtr.grid(row=3, column=0, columnspan=3, sticky="W")
        otpEtr.delete(0, 'end')

        self.dummy(bodyFrame, 4, 0)

        #Back button
        backBtn = ttk.Button(bodyFrame, text="Back", command=lambda: self.loginWidgets(master))
        backBtn.grid(row=5, column=0, sticky="W")

        #Enter button
        enterBtn = ttk.Button(bodyFrame, text="Enter", command=lambda: self.checkPw(master, self.etrVal(otpEtr), noOfAttempt, password=otp, otp=True))
        enterBtn.grid(row=5, column=2, sticky="E")

        master.bind('<Return>', lambda event: self.checkPw(master, self.etrVal(otpEtr), noOfAttempt, password=otp, otp=True, event=True))
   
    def __connectDb__(self, dbPath: str):
        c = sqlite3.connect(dbPath)
        return c

    def checkPw(self, parent, enteredPw, attempts, password="", username="", event=None, otp=False):
        connection = self.__connectDb__(self.__database__)
        cursor = connection.cursor()
        cursor.execute("SELECT * FROM sqlite_master")
        if otp:   #Check OTP
            try:
                # Check whether number of attempts is less than 4
                if attempts < 4:
                    if password == enteredPw:
                        self.loginSuccess = [True]
                        top = tk.Toplevel(root)   #Call new window for main program
                        top.title("Stock Property Management System")
                        width = root.winfo_screenwidth()
                        height = root.winfo_screenheight()   
                        top.geometry(f'{width}x{height}')   #Enlarge window to screen size
                        layout = Layout(top, width, height, login.username, header=True)   #Run main program
                        parent.withdraw()   #Hide root window
                    else:
                        # Increment attempts by 1
                        attempts += 1
                        self.otpWindow(parent, attempts, password, error=True)   #Allow reentry of OTP
                else:
                    self.fail(parent, self.username)
            except:
                self.fail(parent, self.username)

        else:   #Check password
            try:
                self.username = username
                cursor.execute("SELECT * FROM Staff;")
                cursor.execute("SELECT Lock, Pw FROM Staff WHERE Username=:username", {"username": username})   #Check whether the user is locked
                temp = cursor.fetchall()
                if temp == []:
                    self.loginWidgets(parent, attempts=attempts+1, error=True)   #Attempt increments by 1 is error
                lock = temp[0][0]
                password=temp[0][1]
                if attempts <= 4 and lock == False:   #Check whether the user have been or should be locked
                    if password == enteredPw:   #Correct password entered
                        cursor.execute("SELECT Email FROM Staff WHERE Username=:username", {"username": username})
                        email = cursor.fetchall()[0][0]
                        self.otp(parent, self.frames[0], self.frames[1], email)   #Generate and send OTP to user
                    else:
                        attempts += 1
                        if attempts <= 5:
                            self.loginWidgets(parent, attempts=attempts, error=True)   #Return to login page with attempt incremented by 1
                        else:
                            self.fail(parent, self.username)
                else:
                    self.fail(parent, self.username)
            except:
                attempts += 1
                if attempts <= 5:
                    self.loginWidgets(parent, attempts=attempts, error=True)   #If username is wrong attempts still have to increment
                else:
                    self.fail(parent, self.username)

    #Page for locked accounts
    def fail(self, parent, username):
        self.clear(parent)
        outputLbl = ttk.Label(text="""ACCOUNT LOCKED
Please notify Admin
Name: Sunscreen Admin
Pronouns: they/them
Tel no.: 12345678""")
        outputLbl.pack()
        connection = self.__connectDb__(self.__database__)
        cursor = connection.cursor()
        try:
            cursor.execute("UPDATE Staff SET Lock=TRUE WHERE Username=:username", {"username": username})
            connection.commit()
        except:
            pass

    #Send email
    def email(self, userEmail, subject, body):
        companyEmail = 'company.sunscreeninternational@gmail.com'   #Company email 
        emailPw =  'rend jofo kaui sxvc'  #app password generated on myaccount.google.com/apppasswords

        emsg = EmailMessage()
        emsg['From'] = companyEmail
        emsg['To'] = userEmail
        emsg['subject'] = subject
        emsg.set_content(body)

        context = ssl.create_default_context()
        with smtplib.SMTP_SSL('smtp.gmail.com', 465, context=context) as smtp:
            smtp.login(companyEmail, emailPw)
            smtp.sendmail(companyEmail, userEmail, emsg.as_string())
        
class Layout:
    def __init__(self, master, width, height, username, header=True):
        s = ttk.Style()
        s.configure('TFrame', background="light blue")   #Make the background of the frames light blue
        self.__database__ = "./alldata.db"
        connection = self.__connectDb__(self.__database__)   #Connect to database
        cursor = connection.cursor()
        self.clear(master)   #Clear screen
        self.frame = []   #Store frames
        self.font = font.nametofont("TkDefaultFont")
        self.width = width
        self.height = height
        self.page = "Dashboard"   #Landing page name
        self.username = username
        cursor.execute("SELECT StaffFirstName, StaffLastName, AccessLvl FROM Staff WHERE Username=:username", {"username": username})
        staffname = cursor.fetchall()[0]
        self.name = staffname[0]+' '+staffname[1]   #Show staff name on top right corner
        self.accessLvl = staffname[2]   #Get staff access level
        if header:
            self.header(master, self.name, width, height)
        self.menuButton(master, self.page)

        #View purchase order on dashboard when POID clicked
        def view(ID):
            cursor.execute("SELECT SupplierID, PODate, Approval FROM PurchaseOrder WHERE POID=:poid;", {"poid": ID})
            poArr = cursor.fetchall()
            for row in poArr:
                suppID = row[0]
                poDate = row[1]
                approval = row[2]
                if approval != "Rejected" and approval != "Pending":
                    approval = "Approved"
                cursor.execute("SELECT SupplierName FROM Supplier WHERE SupplierID=:suppID", {"suppID": suppID})
                supplierName = cursor.fetchall()[0][0]
                cursor.execute("SELECT SKUID, Qty, TtlPurchasePrice, ExpArrDate, ArrStatus FROM POSKU WHERE POID=:poid", {"poid": ID})
                appendArr = []   #create an array to hold values to be appended into widgetsArr later
                
                #Add PO information to temp array which is then added into appendArr
                for row in cursor.fetchall():
                    temp = []
                    temp.append(row[0])
                    cursor.execute("SELECT SKUName FROM SKU WHERE SKUID=:skuid", {"skuid": row[0]})
                    skuName = cursor.fetchall()[0][0]
                    temp.append(skuName)
                    temp.append(row[1])
                    temp.append(row[2])
                    temp.append(row[3])
                    temp.append(row[4])
                    appendArr.append(temp)

            #Array for all widgets to be shown on the window
            widgetsArr = [[[ttk.Label(master, text="Order No: "), 0, 1, 0, 1, 0], [ttk.Label(master, text=ID), 0, 1, 1, 1, 0], [ttk.Label(master, text="Supplier"), 0, 1, 2, 1, 0], [ttk.Label(master, text=supplierName), 0, 1, 3, 1, 0], [ttk.Label(master, text="PO date"), 0, 1, 4, 1, 0], [ttk.Label(master, text=poDate), 0, 1, 5, 1, 0], [ttk.Label(master, text="Approval: "), 0, 1, 6, 1, 0], [ttk.Label(master, text=approval, foreground="red"), 0, 1, 7, 1, 0], 
                           [ttk.Label(master, text="SKUID"), 1, 1, 0, 1, 0], [ttk.Label(master, text="SKU Name"), 1, 1, 1, 1, 0], [ttk.Label(master, text="QTY"), 1, 1, 2, 1, 0], [ttk.Label(master, text="Total Price"), 1, 1, 3, 1, 0], [ttk.Label(master, text="ETA"), 1, 1, 4, 1, 0], [ttk.Label(master, text="Arrival status"), 1, 1, 5, 1, 0]]]

            rowindex = 2
            #Append the widgets in appendArr into widgetsArr
            for row in appendArr:
                colindex = 0
                for col in row:
                    widgetsArr[0].append([ttk.Label(master, text=col), rowindex, 1, colindex, 1, 0])
                    colindex += 1
                rowindex += 1

            #Save data in database
            def saveDb(poid):
                cursor.execute("UPDATE PurchaseOrder SET Done=TRUE WHERE POID=:poid;", {"poid": poid})
                cursor.execute("SELECT Approval FROM PurchaseOrder WHERE POID=:poid;", {"poid": poid})
                if cursor.fetchall()[0][0] != "Rejected":
                    cursor.execute("SELECT SKUID, Qty FROM POSKU WHERE POID=:poid", {"poid": poid})
                    for rows in cursor.fetchall():
                        cursor.execute("SELECT CurrentStockLvl FROM SKU WHERE SKUID=:skuid", {"skuid": rows[0]})
                        cursor.execute("UPDATE SKU SET CurrentStockLvl=:curr WHERE SKUID=:skuid", {"curr": rows[1]+cursor.fetchall()[0][0], "skuid": rows[0]})
                connection.commit()
                self.__init__(master, width, height, username)   #Return to dashboard
            
            widgetsArr[0].append([ttk.Button(master, text="Done", command=lambda: saveDb(ID)), rowindex, 1, 7, 1, 0])

            self.Content(self, master, widgetsArr, height, width, 0, subframe=1)

        #Widgets on dashboard
        self.widgetsArr = [[[ttk.Label(master, text="Out of Stock", anchor="center", background="light blue"), 0, 1, 0, 4, (width/2-self.font.measure("Out of Stock"))/2], 
                                [ttk.Label(master, text="SKUID", background="light blue"), 1, 1, 0, 1, 0], [ttk.Label(master, text="Current Qty", background="light blue"), 1, 1, 1, 1, 0], [ttk.Label(master, text="Suggested Pur.", background="light blue"), 1, 1, 2, 1, 0], [ttk.Label(master, text="Status", background="light blue"), 1, 1, 3, 1, 0], [ttk.Label(master, text="", background="light blue"), 2, 1, 0, 4, 0]]]

        if self.accessLvl == "staff":
            self.widgetsArr.append( 
                               [[ttk.Label(master, text="Approval/Rejection", anchor="center", background="MediumPurple1", foreground="white"), 0, 1, 0, 3, (width/2-self.font.measure("Approval/Rejection"))/2], 
                                [ttk.Label(master, text="Order No.", background="MediumPurple1", foreground="white"), 1, 1, 0, 1, 0], [ttk.Label(master, text="A/R", background="MediumPurple1", foreground="white"), 1, 1, 1, 1, 0], [ttk.Label(master, text="", background="MediumPurple1"), 2, 1, 0, 2, 0]]
                                )
            #Show approved or rejected POs
            cursor.execute("SELECT POID, Approval FROM PurchaseOrder WHERE DONE=False AND Approval!='Pending'")
            rowindex = 3
            poArr = cursor.fetchall()
            for row in poArr:
                colindex = 0
                for col in row:
                    if colindex == 0:
                        self.widgetsArr[1].append([ttk.Button(master, text=col, cursor="hand2", command=lambda rowindex=rowindex: view(poArr[rowindex-3][0])), rowindex, 1, 0, 1, 0])
                    else:
                        self.widgetsArr[1].append([ttk.Label(master, text=col, background="MediumPurple1", foreground="white"), rowindex, 1, colindex, 1, 0])
                    colindex += 1
                rowindex += 1

        elif self.accessLvl == "manager":
            self.widgetsArr.append( 
                               [[ttk.Label(master, text="To Be Approved/Rejected", anchor="center", background="MediumPurple1", foreground="white"), 0, 1, 0, 3, (width/2-self.font.measure("To Be Approved/Rejected"))/2], 
                                [ttk.Label(master, text="Order No.", background="MediumPurple1", foreground="white"), 1, 1, 0, 1, 0], [ttk.Label(master, text="Vendor Name", background="MediumPurple1", foreground="white"), 1, 1, 1, 1, 0], [ttk.Label(master, text="PO Date", background="MediumPurple1", foreground="white"), 1, 1, 2, 1, 0], [ttk.Label(master, text="", background="MediumPurple1"), 2, 1, 0, 3, 0]]
                                )
            #Select all POs that are not approved or rejected yet
            cursor.execute("SELECT POID, SupplierID, PODate FROM PurchaseOrder WHERE Approval='Pending'")
            rowindex = 3
            poArr = cursor.fetchall()
            for row in poArr:
                colindex = 0
                for col in row:
                    if colindex == 0:
                        self.widgetsArr[1].append([ttk.Button(master, text=col, cursor="hand2", command=lambda rowindex=rowindex: self.approval(master, poArr[rowindex-3][0])), rowindex, 1, 0, 1, 0])
                    else:
                        if colindex == 1:
                            cursor.execute("SELECT SupplierName FROM Supplier WHERE SupplierID=:suppID", {"suppID": col})
                            col = cursor.fetchall()[0][0]
                        self.widgetsArr[1].append([ttk.Label(master, text=col, background="MediumPurple1", foreground="white"), rowindex, 1, colindex, 1, 0])
                    colindex += 1
                rowindex += 1

        #Predict purchase quantity
        self.nn(master)

        dashboard = self.Content(self, master, self.widgetsArr, height, width, 0, subframe = 2)
    
    def __connectDb__(self, dbPath: str):
        c = sqlite3.connect(dbPath)
        return c

    def approval(self, master, poid):
        connection = self.__connectDb__(self.__database__)   #Connect to database
        cursor = connection.cursor()

        #Change page name to 'Approval'
        for widgets in self.frame[1].winfo_children():
            if widgets.grid_info()["column"] == 0 and widgets.grid_info()["row"] == 0:
                widgets.configure({"text": 'Approval'})

        #Details of the PO
        cursor.execute("SELECT * FROM PurchaseOrder WHERE POID=:poid;", {"poid": poid})
        poArr = cursor.fetchall()[0]
        cursor.execute("SELECT SupplierName FROM Supplier WHERE SupplierID=:suppID;", {"suppID": poArr[1]})
        widgetsArr = [[[ttk.Label(master, text="PO No: ", background="light blue"), 0, 1, 0, 1, 0], [ttk.Label(master, text=poArr[0], background="light blue"), 0, 1, 1, 1, 0], [ttk.Label(master, text="Vendor: ", background="light blue"), 0, 1, 2, 1, 0], [ttk.Label(master, text=cursor.fetchall()[0][0], background="light blue"), 0, 1, 3, 1, 0], [ttk.Label(master, text="PODate: ", background="light blue"), 0, 1, 4, 1, 0], [ttk.Label(master, text=poArr[2], background="light blue"), 0, 1, 5, 1, 0], 
                       [ttk.Label(master, text="SKUID", background="light blue"), 1, 1, 0, 1, 0], [ttk.Label(master, text="SKU Name", background="light blue"), 1, 1, 1, 1, 0], [ttk.Label(master, text="QTY", background="light blue"), 1, 1, 2, 1, 0], [ttk.Label(master, text="Total Price", background="light blue"), 1, 1, 3, 1, 0], [ttk.Label(master, text="ETA", background="light blue"), 1, 1, 4, 1, 0]]]
                    
        cursor.execute("SELECT SKUID, Qty, TtlPurchasePrice, ExpArrDate FROM POSKU WHERE POID=:poid", {"poid": poid})
        #Iterate through all the SKUs of the PO
        rowindex = 2
        for row in cursor.fetchall():   
            colindex = 0
            for col in row:
                widgetsArr[0].append([ttk.Label(master, text=col), rowindex, 1, colindex, 1, 0])
                if colindex == 0:   #Add SKU Name into the array
                    colindex += 1
                    cursor.execute("SELECT SKUName FROM SKU WHERE SKUID=:skuid", {"skuid": col})
                    skuname = cursor.fetchall()[0][0]   #Fetch SKU Name
                    widgetsArr[0].append([ttk.Label(master, text=skuname), rowindex, 1, colindex, 1, 0])
                colindex += 1
            rowindex += 1

        #Choose Approve or Reject
        var = tk.StringVar()
        options = ['Approve', 'Reject']
        widgetsArr[0].append([ttk.OptionMenu(master, var, options[0], *options), rowindex, 1, 4, 1, 0])

        #Saving the result into the database
        def saveDb():
            if var.get() == 'Approve':
                entryVal = self.name
            else:
                entryVal = 'Rejected'
            cursor.execute("UPDATE PurchaseOrder SET Approval=:var WHERE POID=:poid", {"var": entryVal, "poid": poid})
            connection.commit()
            self.__init__(master, self.width, self.height, self.username)

        widgetsArr[0].append([ttk.Button(master, text="Save", command=saveDb), rowindex, 1, 5, 1, 0])
        approval = self.Content(self, master, widgetsArr, self.height*0.9, self.width, variable=var)

    def nn(self, master):
        #Fetch all SKUs 
        connection = self.__connectDb__(self.__database__)   #Connect to database
        cursor = connection.cursor()
        cursor.execute("SELECT SKUID FROM SKU;")

        #Load neural network model for each SKU
        urgent = []
        suggested = []
        total = 0
        correct = 0
        for row in cursor.fetchall():
            skuid = row[0]
            if skuid != 'TC-0001' and skuid!= 'TC-0004' and skuid!='TC-0037' and skuid!='TC-0053' and skuid != 'HW-0277' and skuid!='HW-0278' and skuid!='HW-0279' and skuid!='HW-0281':
                newModel = model.Model()
            else:
                newModel = model.Model(inputs=22)
            ptname = skuid + 'nn.pt'
            try:   #Take into account new SKUs that doesn't have a model yet
                newModel.load_state_dict(model.torch.load(ptname))

                preProcess = model.Process(skuid)

                #Get today's date
                today = datetime.date.today()

                todayPlus7 = today + datetime.timedelta(days=7)
                todayStr = today.strftime(r'%d/%m/%Y')
                month = today.strftime('-%m')
                cursor.execute("SELECT VIP, SD1, SD2 FROM SPECIAL WHERE SpecialDate=:date", {"date": todayStr})
                specialArr = cursor.fetchall()[0]

                cursor.execute("SELECT CurrentStockLvl FROM SKU WHERE SKUID=:skuid", {"skuid": skuid})
                currentQty = cursor.fetchall()[0][0]
                
                xInput = [specialArr[0]]
                if skuid == 'TC-0001' or skuid== 'TC-0004' or skuid=='TC-0037' or skuid=='TC-0053':
                    xInput.append(specialArr[1])
                if skuid == 'HW-0277' or skuid=='HW-0278' or skuid=='HW-0279' or skuid=='HW-0281':
                    xInput.append(specialArr[2])
                #OneHotEncode months
                for j in range (12):
                    if j == int(month)-1:
                        xInput.append(1)
                    else:
                        xInput.append(0)
                #OneHotEncode day of week
                for j in range (7):
                    if j == today.isoweekday()-1:
                        xInput.append(1)
                    else:
                        xInput.append(0)

                #Standardise quantity
                xInput.append((currentQty-preProcess.qtyMean)/preProcess.qtyStd)

                xInput = model.torch.FloatTensor(xInput)
                yOutput = newModel.forwardPass(xInput)
                yOutput = round(float(yOutput.item())*preProcess.qty7Std+preProcess.qty7Mean, 1)

                #Fetch all POSKU with expected arrival date before 7 days later
                cursor.execute("SELECT ExpArrDate, Qty FROM POSKU WHERE SKUID!=skuid AND ArrStatus='Not Arrived';", {"skuid": skuid})
                poskuarr = cursor.fetchall()
                qty7 = currentQty

                #list of SKUs that need to be shown
                for thing in poskuarr:
                    if datetime.strptime(thing[0], r'%d/%m/%Y') < todayPlus7:
                        qty7 += thing[1]
                    #If quantity after 7 days is negative, add quantity up to qrange*0.1 (base stock)
                
                yOutput = 1.1*yOutput   #Counter underprediction as accuracy is based on 10% within actual quantity
                if qty7 <= 0:   #If there's no more stock
                    urgent.append([skuid, qty7, round(preProcess.qrange*0.1-qty7, 1)])
                elif yOutput-qty7 > yOutput*0.1:   #If difference > 10% of predicted
                    urgent.append([skuid, qty7, round(yOutput-qty7, 1)])
                elif qty7 < yOutput and yOutput-qty7 <= yOutput*0.1:   #If difference < 10% of predicted
                    suggested.append([skuid, qty7, round(yOutput-qty7, 1)])

            except:
                pass

        #Show urgent and suggested SKUs on window
        rowindex = 3
        for info in urgent:
            self.widgetsArr[0].append([ttk.Label(master, text=info[0]), rowindex, 1, 0, 1, 0])
            self.widgetsArr[0].append([ttk.Label(master, text=info[1]), rowindex, 1, 1, 1, 0])
            self.widgetsArr[0].append([ttk.Label(master, text=info[2]), rowindex, 1, 2, 1, 0])
            self.widgetsArr[0].append([ttk.Label(master, text="Urgent", foreground="red"), rowindex, 1, 3, 1, 0])
            rowindex += 1
        for info in suggested:
            self.widgetsArr[0].append([ttk.Label(master, text=info[0]), rowindex, 1, 0, 1, 0])
            self.widgetsArr[0].append([ttk.Label(master, text=info[1]), rowindex, 1, 1, 1, 0])
            self.widgetsArr[0].append([ttk.Label(master, text=info[2]), rowindex, 1, 2, 1, 0])
            self.widgetsArr[0].append([ttk.Label(master, text="Suggested", foreground="light blue"), rowindex, 1, 3, 1, 0])
            rowindex += 1

    #Clear frame
    def clear(self, frame):
        for widget in frame.winfo_children():
            widget.destroy()

    def header(self, master, staffName, width, height):
        #Company name which is a button
        s = ttk.Style(master)
        s.configure('Frame0.TFrame', background='light blue')
        s.configure('TLabel', background='light blue', foreground="black")
        s.layout("TButton", [('Button.button', {'sticky': 'nswe', 'children': [('Button.focus', {'sticky': 'nswe', 'children': [('Button.padding', {'sticky': 'nswe', 'children': [('Button.label',
    {'sticky': 'w'})]})]})]})]) #To allow the name of the company to stick to the far left
        
        #Header
        headerFrame = ttk.Frame(master, style='TFrame', width=width, height=height*0.05)
        self.frame.append(headerFrame)
        headerFrame.grid_propagate(False)
        headerFrame.grid(row=0)
        titleBtn = ttk.Button(headerFrame, text="太陽油集團（國際）有限公司", cursor="hand2", command=lambda: self.__init__(master, width, height, self.username, header=True))
        titleBtn.grid(row=0, column=0, sticky="w")
        titleBtn.grid_propagate(0)

        #Logout
        self.variable = tk.StringVar(headerFrame)
        options = [staffName, 'Log out']
        def click(value):
            if value == 'Log out':
                master.master.destroy()   #Destroy root window
        profileBtn = ttk.OptionMenu(headerFrame, self.variable, options[0], options[1], command=click)
        profileBtn.grid(row=0, column=2)
        progNameLbl = ttk.Label(headerFrame, text="Stock Management System", anchor=tk.CENTER)
        master.update_idletasks() 
        txtName = progNameLbl.cget("text")
        txtWidth = self.font.measure(txtName)   #Make program name stay in the middle
        progNameLbl.grid(row=0, column=1, ipadx=((width-titleBtn.winfo_width()-profileBtn.winfo_width()-txtWidth)//2))

    def menuButton(self, master, page, accessLvl="manager"):
        #Create frame for menu
        menuFrame = ttk.Frame(master, width=self.width, height=self.height*0.05)
        menuFrame.grid(row=1)
        self.frame.append(menuFrame)

        s = ttk.Style(master)
        s.configure('TLabel', background='light blue', foreground="black")
        s.configure('TButton', foreground='grey', cursor="hand2")
        s.layout("TButton", [('Button.button', {'sticky': 'nswe', 'children': [('Button.focus', {'sticky': 'nswe', 'children': [('Button.padding', {'sticky': 'nswe', 'children': [('Button.label',
    {'sticky': 'w'})]})]})]})])
            
        #Page name
        pageTxt = page
        pageLbl = ttk.Label(menuFrame, text=pageTxt)
        pageLbl.grid(row=0, column=0)

        #Menu Button
        def btnClick(cls, master, root, index):
            connection = self.__connectDb__(self.__database__)   #Connect to database
            cursor = connection.cursor()
            
            #Destroy original page
            for widget in cls.frame[2].winfo_children():
                    widget.destroy()
            
            #New PO page
            def addPurchase(cls, root):
                    #Get today's year
                    today = datetime.date.today()
                    year = today.year

                    #Find last POID and add 1
                    connection = self.__connectDb__(self.__database__)   #Connect to database
                    cursor = connection.cursor()
                    cursor.execute("SELECT POID FROM PurchaseOrder ORDER BY POID DESC LIMIT 1;")
                    lastPO = cursor.fetchall()[0][0]

                    if int(lastPO[:4]) != year:
                        num = 1
                    else:
                        num = int(lastPO[-4:])+1
                    if num<10:
                        poNo = str(year)+"-000"+str(num)
                    elif num<100:
                        poNo = str(year)+"-00"+str(num)
                    elif num<1000:
                        poNo = str(year)+"-0"+str(num)
                    else:
                        poNo = str(year)+"-"+str(num)

                    #Supplier option list
                    variable = tk.StringVar(master, "--Select--")
                    supplierOptn = []
                    #Search through Supplier table
                    cursor.execute("Select SupplierName from Supplier")
                    for row in cursor.fetchall():
                        supplierOptn.append(row[0])   #add all supplier names into the Optionmenu
                    
                    poDate = tk.StringVar()

                    #Show correct widget on correct coordinates
                    def show(rowindex):
                        for widget in newPO.frames[0].winfo_children():
                            try:
                                if widget.grid_info()["column"] == 0 and widget.grid_info()["row"] == rowindex:
                                    cursor.execute("SELECT SKUName FROM SKU WHERE SKUID=:skuid", {"skuid": widget.get()})
                                    skuname = cursor.fetchall()[0][0]
                                elif widget.grid_info()["column"] == 1 and widget.grid_info()["row"] == rowindex:
                                    widget.configure({"text": skuname, "foreground": "black"})
                            except:
                                if widget.grid_info()["column"] == 1 and widget.grid_info()["row"] == rowindex:
                                    widget.configure({"text": "Invalid SKUID", "foreground": "red"})

                    #Add new row to input products
                    def addRow(self, array, rowindex):
                        #Destroy previous "+" button and "Send to Approval" button
                        for widget in newPO.frames[0].winfo_children():
                            try:
                                if widget.cget("text") == "+" or widget.cget("text") == "Send to Approval":
                                    widget.destroy()
                            except:
                                pass

                        wholeRow = [[ttk.Label(newPO.frames[0], text="SKUID"), rowindex, 1, 0, 1, 0], [ttk.Button(newPO.frames[0], text="Enter", command=lambda rowindex=rowindex: show(rowindex+1)), rowindex, 1, 1, 1, 0], [ttk.Label(newPO.frames[0], text="QTY"), rowindex, 1, 2, 1, 0], [ttk.Label(newPO.frames[0], text="Total Price"), rowindex, 1, 3, 1, 0], [ttk.Label(newPO.frames[0], text="ETA"), rowindex, 1, 4, 1, 0], 
                                    [ttk.Entry(newPO.frames[0]), rowindex+1, 1, 0, 1, 0], [ttk.Label(newPO.frames[0]), rowindex+1, 1, 1, 1, 0], [ttk.Entry(newPO.frames[0]), rowindex+1, 1, 2, 1, 0], [ttk.Entry(newPO.frames[0]), rowindex+1, 1, 3, 1, 0], [ttk.Entry(newPO.frames[0]), rowindex+1, 1, 4, 1, 0], 
                                    [ttk.Button(newPO.frames[0], text="+", command=lambda: addRow(self, array, rowindex+2)), rowindex+2, 1, 0, 1, 0],
                                    [ttk.Button(newPO.frames[0], text="Send to Approval", command=saveDb), rowindex+3, 1, 4, 1, 0]]

                        for j in wholeRow:
                            j[0].grid(row=j[1], rowspan=j[2], column=j[3], columnspan=j[4], ipadx=j[5])
                            array.append(j)
                        
                        #Update scrollbar region
                        newPO.canvasFrame.bind('<Configure>', lambda e: newPO.canvas.config(scrollregion=newPO.canvas.bbox("all")))

                    def saveDb():
                        entryVal = []
                        connection = self.__connectDb__(self.__database__)   #Connect to database
                        cursor = connection.cursor()

                        #Get entered values
                        for widget in newPO.frames[0].winfo_children():
                            if widget.grid_info()["column"] == 1 and widget.grid_info()["row"] == 0:
                                POID = widget.cget("text")
                            elif widget.winfo_class() == "TEntry":
                                entryVal.append(widget.get())
                            elif widget.winfo_class() == "TMenubutton":
                                supplier = variable.get()

                        #Update PurchaseOrder through sql lines
                        cursor.execute("SELECT SupplierID FROM Supplier WHERE SupplierName=:sup;", {"sup": supplier})
                        supplierID = cursor.fetchall()[0][0]
                        
                        #Check format of POdate
                        try:
                            poDate = datetime.datetime.strptime(entryVal[0], r'%d/%m/%Y')
                            cursor.execute("INSERT INTO PurchaseOrder(POID, SupplierID, PODate, Approval, Done) VALUES(?, ?, ?, ?, ?);", (POID, supplierID, entryVal[0], "Pending", False))

                            for j in range ((len(entryVal)-1)//4):
                                #Check format of Qty, ETA and Total price
                                qty=float(entryVal[4*j+2])
                                price=float(entryVal[4*j+3])
                                expDate = datetime.datetime.strptime(entryVal[4*j+4], r'%d/%m/%Y')
                                cursor.execute("SELECT POSKUID FROM POSKU ORDER BY POSKUID DESC LIMIT 1;")
                                POSKUID = cursor.fetchall()[0][0]+1
                                cursor.execute("INSERT INTO POSKU VALUES(?, ?, ?, ?, ?, ?, ?);", (POSKUID, POID, entryVal[4*j+1], qty, price, entryVal[4*j+4], 'Not Arrived'))

                            connection.commit()
                            connection.close()
                            addPurchase(cls, root)

                        except:
                            ttk.Label(newPO.frames[0], text="Invalid Entry", foreground="red").grid(row=0, column=10)

                    widgetsArr = [[[ttk.Label(master, text="Order No: "), 0, 1, 0, 1, 0], [ttk.Label(master, text=poNo), 0, 1, 1, 1, 0], [ttk.Label(master, text="Supplier"), 0, 1, 2, 1, 0], [ttk.Label(master, text="PO date"), 0, 1, 3, 1, 0], 
                                   [ttk.OptionMenu(master, variable, *supplierOptn), 1, 1, 2, 1, 0], [ttk.Entry(master, textvariable=poDate), 1, 1, 3, 1, 0],
                                   [ttk.Label(master, text="SKUID"), 2, 1, 0, 1, 0], [ttk.Button(master, text="Enter", command=lambda: show(3)), 2, 1, 1, 1, 0], [ttk.Label(master, text="QTY"), 2, 1, 2, 1, 0], [ttk.Label(master, text="Total Price"), 2, 1, 3, 1, 0], [ttk.Label(master, text="ETA"), 2, 1, 4, 1, 0], 
                                   [ttk.Entry(master), 3, 1, 0, 1, 0], [ttk.Label(master), 3, 1, 1, 1, 0], [ttk.Entry(master), 3, 1, 2, 1, 0], [ttk.Entry(master), 3, 1, 3, 1, 0], [ttk.Entry(master), 3, 1, 4, 1, 0], 
                                   [ttk.Button(master, text="+", command=lambda: addRow(self, widgetsArr, 4)), 4, 1, 0, 1, 0],
                                   [ttk.Button(master, text="Send to Approval", command=saveDb), 5, 1, 4, 1, 0]]]
                    newPO = cls.Content(cls, root, widgetsArr, cls.height*0.9, cls.width, variable)

            #New GRN page
            def addReceive(cls, root):
                    def saveDb(grnNo, poskuid, actDate):
                        try:
                            #Validate the date
                            actdate = datetime.datetime.strptime(actDate, r'%d/%m/%Y')
                            
                            #Update database
                            cursor.execute("UPDATE POSKU SET ArrStatus=:arrStat WHERE POSKUID=:poskuid;", {"arrStat": "Arrived", "poskuid": poskuid})
                            cursor.execute("INSERT INTO GRN (GRNNo, POSKUID, ActArrDate) VALUES (?, ?, ?);", (grnNo, poskuid, actDate))
                            cursor.execute("SELECT SKUID, Qty FROM POSKU WHERE POSKUID=:poskuid", {"poskuid": poskuid})
                            poskuarr = cursor.fetchall()[0]
                            cursor.execute("SELECT CurrentStockLvl FROM SKU WHERE SKUID=:skuid", {"skuid": poskuarr[0]})
                            newCurrent = cursor.fetchall()[0][0]+poskuarr[1]
                            cursor.execute("UPDATE SKU SET CurrentStockLvl=:newCurrent WHERE SKUID=:skuid", {"newCurrent": newCurrent, "skuid": poskuarr[0]})
                            connection.commit()
                            addReceive(cls, root)   #Allow second entry immediately
                        except:
                            ttk.Label(receivePg.frames[0], text="Invalid Entry", foreground="red").grid(row=0, column=8)

                    #Fetch record in POSKU according to POID and SKUID entered
                    def find():
                        secRow = []
                        entryVal = []
                        #Retrieve PONumber and SKUID
                        for widget in receivePg.frames[0].winfo_children():
                            if widget.winfo_class() == "TEntry":
                                entryVal.append(widget.get())

                        #Fetch quantity, total purchase price, expected arrival date and sku name
                        try:   #Prevent wrong SKUID or PO
                            cursor.execute("SELECT POSKUID, Qty, TtlPurchasePrice, ExpArrDate FROM POSKU WHERE POID=:poid AND SKUID=:skuid", {"poid": entryVal[0], "skuid": entryVal[1]})
                            info = cursor.fetchall()
                            poskuid, qty, total, expDate = info[0][0], info[0][1], info[0][2], info[0][3]
                            cursor.execute("SELECT SKUName FROM SKU WHERE SKUID=:skuid", {"skuid": entryVal[1]})
                            skuName = cursor.fetchall()[0][0]
                            cursor.execute("SELECT SupplierID, PODate FROM PurchaseOrder WHERE POID=:poid", {"poid": entryVal[0]})
                            info2 = cursor.fetchall()
                            podate = info2[0][1]
                            cursor.execute("SELECT SupplierName FROM Supplier WHERE SupplierID=:suppID", {"suppID": info2[0][0]})
                            suppname = cursor.fetchall()[0][0]
                            secRow.append(entryVal[1])
                            secRow.append(skuName)
                            secRow.append(qty)
                            secRow.append(total)
                            secRow.append(expDate)
                            col = 0
                            for info in secRow:
                                colspan = 1
                                if col == 0:
                                    colspan = 2
                                ttk.Label(receivePg.frames[0], text=info).grid(row=2, rowspan=1, column=col+2, columnspan=colspan)   #Print info about the SKU in the PO
                                col += 1
                            actDate = ttk.Entry(receivePg.frames[0])
                            actDate.grid(row=2, rowspan=1, column=col+2, columnspan=1)   #entry of actual arrival date
                            ttk.Button(receivePg.frames[0], text="Save", command=lambda: saveDb(grnNo, poskuid, actDate.get())).grid(row=3, rowspan=1, column=col+2, columnspan=1)   #Save into database

                            #Update supplier name and PODate label
                            labels = []
                            for widget in receivePg.frames[0].winfo_children():
                                if widget.winfo_class() == 'TLabel' and widget.cget("text") == "":
                                    labels.append(widget)
                                if widget.grid_info()["column"] == 8:
                                    widget.destroy()
                            labels[0].configure(text=suppname)
                            labels[1].configure(text=podate)


                        except:   #Show error label for invalid entry and destroy existing information
                            ttk.Label(receivePg.frames[0], text="Invalid Entry", foreground="Red").grid(row=0, column=8) 
                            for widget in receivePg.frames[0].winfo_children():
                                if widget.grid_info()["row"] >= 2:
                                    widget.destroy()

                    #Fetch last GRNNo and calculate new GRNNo
                    cursor.execute("SELECT GRNNo FROM GRN ORDER BY GRNNo DESC LIMIT 1;")
                    currGRNNo = cursor.fetchall()[0][0]
                    GRNYr = currGRNNo[4:8]
                    today = datetime.date.today()
                    if today >= datetime.date(today.year, 4, 1):
                        yr = str(today.year)[2:]
                        if yr[:1] == '0':
                            yr == yr[1:]
                        term = yr + str(int(yr)+1)
                    else:
                        yr = str(today.year)[2:]
                        if yr[:1] == '0':
                            yr == yr[1:]
                        term = str(int(yr)-1) + yr
                    
                    if term == GRNYr:
                        num = currGRNNo[9:]
                        num = num.strip()
                        num = num.replace('0', '') 
                        num = int(num)+1
                        if num <= 9:
                            grnNo = 'GRN-' + term + '-0000' + str(num)
                        elif num <= 99:
                            grnNo = 'GRN-' + term + '-000' + str(num)
                        elif num <= 999:
                            grnNo = 'GRN-' + term + '-00' + str(num)
                        elif num <= 9999:
                            grnNo = 'GRN-' + term + '-0' + str(num)
                        else:
                            num = 'GRN-' + term + '-' + str(num)
                    else:
                        grnNo = 'GRN-' + term + '-' + '00001'

                    #Create widgets
                    widgetsArr = [[[ttk.Label(master, text="PO No: "), 0, 1, 0, 1, 0], [ttk.Entry(master), 0, 1, 1, 1, 0], [ttk.Label(master, text="GRN: "), 0, 1, 2, 1, 0], [ttk.Label(master, text=grnNo), 0, 1, 3, 1, 0], [ttk.Label(master, text="Supplier: "), 0, 1, 4, 1, 0], [ttk.Label(master), 0, 1, 5, 1, 0], [ttk.Label(master, text="PO date: "), 0, 1, 6, 1, 0], [ttk.Label(master), 0, 1, 7, 1, 0], 
                                   [ttk.Label(master, text="SKUID"), 1, 1, 0, 1, 0], [ttk.Entry(master), 1, 1, 1, 1, 0], [ttk.Button(master, text="Enter", command=find), 1, 1, 2, 1, 0], [ttk.Label(master, text="SKU Name"), 1, 1, 3, 1, 0], [ttk.Label(master, text="Quantity"), 1, 1, 4, 1, 0], [ttk.Label(master, text="Total Purchase Price"), 1, 1, 5, 1, 0], [ttk.Label(master, text="Expected Arrival Date"), 1, 1, 6, 1, 0], [ttk.Label(master, text="Actual Arrival Date"), 1, 1, 7, 1, 0], 
                                   ]]
                    
                    receivePg = cls.Content(cls, root, widgetsArr, cls.height*0.9, cls.width, 0)

            def addSales(cls, root):
                    #Open daily order excel file
                    filepath = filedialog.askopenfilename()   #Allow users to choose file
                    df = openpyxl.load_workbook(filepath)
                    df1 = df.active

                    for row in range(1, df1.max_row+1):   #iterate through all rows in the excel file
                        if row >= 6:
                            soid = df1.cell(row=row, column=4).value
                            soDate = df1.cell(row=row, column=7).value
                            soDate = str(soDate)
                            soDate = soDate[8:] + "/" + soDate[5:7] + "/" + soDate[:4] 
                            dvDate = df1.cell(row=row, column=11).value
                            dvDate = str(dvDate)
                            dvDate = dvDate[8:] + "/" + dvDate[5:7] + "/" + dvDate[:4]
                            skuid = df1.cell(row=row, column=18).value
                            qty = df1.cell(row=row, column=24).value
                            total = df1.cell(row=row, column=27).value
                            dvStatus = df1.cell(row=row, column=28).value

                            #Update current stock level
                            if dvStatus != 'CONFIRMED':
                                cursor.execute("SELECT CurrentStockLvl FROM SKU WHERE SKUID=:skuid;", {"skuid": skuid})
                                curr = cursor.fetchall()[0][0]
                                curr -= qty
                                cursor.execute("UPDATE SKU SET CurrentStockLvl=:curr WHERE SKUID=:skuid", {"curr": curr, "skuid": skuid})

                            cursor.execute("SELECT SOID FROM SalesOrder WHERE SOID=:soid", {"soid": soid})
                            if cursor.fetchall() != []:   #If there's an SO previously
                                cursor.execute("SELECT SOSKUID FROM SOSKU WHERE SOID=:soid AND SKUID=:skuid;", {"soid": soid, "skuid": skuid})
                                soskuarr = cursor.fetchall()
                                #Check whether the SOSKU record is created before, if not, create a new record, if yes, update the record
                                if soskuarr == []:
                                    cursor.execute("SELECT SOSKUID FROM SOSKU ORDER BY SOSKUID DESC LIMIT 1;")
                                    cursor.execute("INSERT INTO SOSKU VALUES(?, ?, ?, ?, ?, ?, ?)", (cursor.fetchall()[0][0]+1, soid, skuid, qty, total, dvDate, dvStatus))
                                else:
                                    cursor.execute("UPDATE SOSKU SET DeliveryDate=:dvDate, DeliveryStatus=:dvStatus WHERE SOSKUID=:soskuid;", {"dvDate": dvDate, "dvStatus": dvStatus, "soskuid": soskuarr[0][0]})
                            
                            else:   #Create new SO
                                cursor.execute("INSERT INTO SalesOrder VALUES(?, ?)", (soid, soDate))
                                cursor.execute("SELECT SOSKUID FROM SOSKU ORDER BY SOSKUID DESC LIMIT 1;")
                                cursor.execute("INSERT INTO SOSKU VALUES(?, ?, ?, ?, ?, ?, ?)", (cursor.fetchall()[0][0]+1, soid, skuid, qty, total, dvDate, dvStatus))
                        
                            connection.commit()
                            
                    ttk.Label(sales.frames[0], text="Update Successful", style="Error.TLabel").grid(row=0, rowspan=1, column=7, columnspan=1)

            def addVendors(cls, root):
                    connection = self.__connectDb__(self.__database__)   #Connect to database
                    cursor = connection.cursor()

                    #Add new record
                    def saveDb(suppID):
                        for widget in newVendor.frames[0].winfo_children():
                            if widget.winfo_class() == 'TEntry':
                                suppName = (widget.get())
                        cursor.execute("INSERT INTO Supplier VALUES (?, ?);", (suppID, suppName))
                        connection.commit()
                        addVendors(cls, root)
                    
                    #fetch the newest vendor id from db and generate new vendor ID
                    cursor.execute("SELECT SupplierID FROM Supplier ORDER BY SupplierID DESC LIMIT 1;")
                    suppID = cursor.fetchall()[0][0]
                    index = suppID[3:]
                    if index[0] == '0':
                        index.strip()
                        index = index.replace('0', '')
                    index = int(index)+1
                    if index < 10:
                        suppID = 'PRC00' + str(index)
                    elif index <100:
                        suppID = 'PRC0' + str(index)
                    else:
                        suppID = 'PRC' + str(index)

                    #Widgets on the page
                    widgetsArr = [[[ttk.Label(master, text="Vendor ID: "), 0, 1, 0, 1, 0], [ttk.Label(master, text=suppID), 0, 1, 1, 1, 0], 
                                   [ttk.Label(master, text="Vendor name: "), 1, 1, 0, 1, 0], [ttk.Entry(master), 1, 1, 1, 1, 0], 
                                  [ttk.Button(master, text="Enter", command=lambda: saveDb(suppID)), 2, 1, 1, 1, 0]]]
                    
                    newVendor = cls.Content(cls, root, widgetsArr, cls.height*0.9, cls.width, variable=0)

            def addSKU(cls, root):
                    #Save to database
                    def saveDb():
                        entryVal = []
                        for widget in newSKU.frames[0].winfo_children():
                            if widget.winfo_class() == 'TEntry':
                                entryVal.append(widget.get())

                        connection = self.__connectDb__(self.__database__)   #Connect to database
                        cursor = connection.cursor()

                        try:
                            #Check data type of quantity
                            qty = float(entryVal[2])
                            cursor.execute("SELECT SKUID FROM SKU WHERE SKUID=:skuid;", {"skuid": entryVal[0]})
                            if cursor.fetchall() != []:
                                s.configure("Error.TLabel", foreground="red")
                                ttk.Label(newSKU.frames[0], text="Error: SKUID is used", style="Error.TLabel").grid(row=3, rowspan=1, column=0, columnspan=1)
                            else:
                                cursor.execute("INSERT INTO SKU VALUES (?, ?, ?, ?)", (entryVal[0], entryVal[1], qty, qty))
                                connection.commit()
                                addSKU(cls, root)
                        except:
                            #Error label
                            ttk.Label(newSKU.frames[0], text="Invalid Entry", foreground="red").grid(row=3, rowspan=1, column=0, columnspan=1)

                    widgetsArr = [[[ttk.Label(master, text="SKUID: "), 0, 1, 0, 1, 0], [ttk.Entry(master, ), 0, 1, 1, 1, 0], 
                                   [ttk.Label(master, text="SKU name: "), 1, 1, 0, 1, 0], [ttk.Entry(master), 1, 1, 1, 1, 0], 
                                   [ttk.Label(master, text="Opening Stock: "), 2, 1, 0, 1, 0], [ttk.Entry(master), 2, 1, 1, 1, 0], 
                                  [ttk.Button(master, text="Enter", command=saveDb), 3, 1, 1, 1, 0]]]

                    newSKU = cls.Content(cls, root, widgetsArr, cls.height*0.9, cls.width, variable=0)

            def search(page):
                    #link each page to their respective Content object so the frame can be correctly updated
                    match (page):
                        case "PO":
                            content=purchase
                        case "GRN":
                            content=receive
                        case "SO":
                            content=sales
                        case "SKU":
                            content=stock
                    connection = self.__connectDb__(self.__database__)   #Connect to database
                    cursor = connection.cursor()
                    entryVal = []
                    for widget in content.frames[0].winfo_children():
                        if widget.winfo_class() == "TEntry":
                            entryVal.append(widget.get())
                        if widget.grid_info()['row']>=2:
                            widget.destroy()
                    
                    if page != "SKU":
                        #Search according to time period
                        if entryVal[0] != "" and entryVal[1]!= "":
                            startDate = datetime.datetime.strptime(entryVal[0], r"%d/%m/%Y")
                            endDate = datetime.datetime.strptime(entryVal[1], r"%d/%m/%Y")
                            delta = endDate - startDate
                            tempArr = []
                            for j in range (delta.days+1):
                                findDate = startDate + datetime.timedelta(days=j)
                                findDate = findDate.strftime(r"%d/%m/%Y")
                                if page == "PO":
                                    cursor.execute("SELECT * FROM PurchaseOrder WHERE PODate=:date", {"date": findDate})
                                elif page == "GRN":
                                    cursor.execute("SELECT * FROM GRN WHERE ActArrDate=:date", {"date": findDate})
                                elif page == "SO":
                                    cursor.execute("SELECT * FROM SalesOrder WHERE SODate=:date", {"date": findDate})
                                try:
                                    tempArr.append(cursor.fetchall()[0])
                                except:
                                    pass
                                
                            rowindex = 0
                            for row in tempArr:
                                total=0
                                required = []
                                match (page):
                                    case "PO":
                                        #Fetch information
                                        poid=row[0]
                                        cursor.execute("SELECT SupplierName FROM Supplier WHERE SupplierID=:supID", {"supID": row[1]})
                                        suppName = cursor.fetchall()[0][0]
                                        poDate = row[2]
                                        approval = row[3]
                                        if approval!="Rejected" and approval!="Pending":
                                            approval = "Approved by " + approval
                                        total = 0
                                        cursor.execute("SELECT TtlPurchasePrice FROM POSKU WHERE POID=:poid", {"poid": poid})

                                        #Add up price of each POSKU
                                        for amt in cursor.fetchall():
                                            if amt[0].__class__.__name__ == 'str':
                                                adder = amt[0].strip()
                                                adder = adder.replace(',', '') 
                                                adder = float(adder)
                                            else:
                                                adder = amt[0]
                                            total += adder

                                        #Add information into array to be shown on window
                                        required.append(poid)
                                        required.append(suppName)
                                        required.append(poDate)
                                        required.append(approval)
                                        required.append(str(round(total,2)))
                                    case "GRN":
                                        #Fetch information
                                        grnno = row[0]
                                        arrDate = row[2]
                                        cursor.execute("SELECT POID, SKUID, Qty, TtlPurchasePrice FROM POSKU WHERE POSKUID=:poskuid", {"poskuid": row[1]})
                                        fetch = cursor.fetchall()
                                        poid = fetch[0][0]
                                        skuid = fetch[0][1]
                                        qty = fetch[0][2]

                                        #Add into array to be shown onto window
                                        cursor.execute("SELECT SKUName FROM SKU WHERE SKUID=:skuid", {"skuid": skuid})
                                        required.append(grnno)
                                        required.append(skuid)
                                        required.append(cursor.fetchall()[0][0])
                                        required.append(qty)
                                        required.append(arrDate)
                                    case "SO":
                                        #Fetch information
                                        soid = row[0]
                                        soDate = row[1]
                                        cursor.execute("SELECT TtlSellingPrice FROM SOSKU WHERE SOID=:soid", {"soid": soid})
                                        for amt in cursor.fetchall():
                                            if amt[0].__class__.__name__ == 'str':
                                                adder = amt[0].strip()
                                                adder = adder.replace(',', '') 
                                                adder = float(adder)
                                            else:
                                                adder = amt[0]
                                            total += adder
                                            
                                        #Add into array to be shown on window
                                        required.append(soid)
                                        required.append(soDate)
                                        required.append(str(round(total,2)))
                                        
                                #Put required array onto window
                                for times in range(len(required)):
                                    if times == 0:
                                        ID = required[0]
                                        ttk.Button(content.frames[0], text=ID, command=lambda ID=ID: view(page, ID)).grid(row=rowindex+2, rowspan=1, column=0, columnspan=1)
                                    else:
                                        ttk.Label(content.frames[0], text=required[times]).grid(row=rowindex+2, rowspan=1, column=times, columnspan=1)
                                rowindex += 1
                                    
                        #Search according to PONo/GRNNo/SONo
                        elif entryVal[2]!="":
                            required = []
                            try:
                                match (page):
                                    case"PO":
                                            #Fetch information
                                            cursor.execute("SELECT SupplierID, PODate, Approval FROM PurchaseOrder WHERE POID=:poid", {"poid": entryVal[2]})
                                            infoArr = cursor.fetchall()
                                            cursor.execute("SELECT SupplierName FROM Supplier WHERE Supplierid=:suppID", {"suppID": infoArr[0][0]})
                                            suppName = cursor.fetchall()[0][0]
                                            poDate = infoArr[0][1]
                                            approval = infoArr[0][2]

                                            if approval!="Rejected" and approval!="Pending":
                                                approval = "Approved by " + approval
                                            cursor.execute("SELECT TtlPurchasePrice FROM POSKU WHERE POID=:poid", {"poid": entryVal[2]})
                                            total = 0

                                            #Add up total amount
                                            for amt in cursor.fetchall():
                                                if amt[0].__class__.__name__ == 'str':
                                                    adder = amt[0].strip()
                                                    adder = adder.replace(',', '') 
                                                    adder = float(adder)
                                                else:
                                                    adder = amt[0]
                                                total += adder  
                                            
                                            #Put widgets into required array
                                            required.append(entryVal[2])
                                            required.append(suppName)
                                            required.append(poDate)
                                            required.append(approval)
                                            required.append(str(round(total, 2)))                                          
                                    case "GRN":
                                            #Fetch information
                                            cursor.execute("SELECT POSKUID, ActArrDate FROM GRN WHERE GRNNo=:grnno", {"grnno": entryVal[2]})
                                            infoArr = cursor.fetchall()[0]
                                            poskuid = infoArr[0]
                                            arrDate = infoArr[1]
                                            cursor.execute("SELECT POID, SKUID, Qty FROM POSKU WHERE POSKUID=:poskuid", {"poskuid": poskuid})
                                            poArr = cursor.fetchall()[0]
                                            poid = poArr[0]
                                            skuid = poArr[1]
                                            qty = poArr[2]
                                            cursor.execute("SELECT SKUName FROM SKU WHERE SKUID=:skuid", {"skuid":skuid})
                                            skuname = cursor.fetchall()[0][0]

                                            #Add onto required array
                                            required.append(entryVal[2])
                                            required.append(skuid)
                                            required.append(skuname)
                                            required.append(qty)
                                            required.append(arrDate)
                                    case "SO": 
                                            #Fetch inforamation
                                            cursor.execute("SELECT SODate FROM SalesOrder WHERE SOID=:soid", {"soid": entryVal[2]})
                                            soDate = cursor.fetchall()[0][0]
                                            cursor.execute("SELECT TtlSellingPrice FROM SOSKU WHERE SOID=:soid", {"soid": entryVal[2]})
                                            total = 0

                                            #Calculate sum of prices of SOSKUs
                                            for amt in cursor.fetchall():
                                                if amt[0].__class__.__name__ == 'str':
                                                    adder = amt[0].strip()
                                                    adder = adder.replace(',', '') 
                                                    adder = float(adder)
                                                else:
                                                    adder = amt[0]
                                                total += adder  

                                            #Add onto required array
                                            required.append(entryVal[2])
                                            required.append(soDate)
                                            required.append(str(round(total, 2)))

                                #Put widgets on window
                                for times in range(len(required)):
                                    if times == 0:
                                        ID = required[0]
                                        ttk.Button(content.frames[0], text=ID, command=lambda: view(page, ID)).grid(row=2, rowspan=1, column=0, columnspan=1)
                                    else:
                                        ttk.Label(content.frames[0], text=required[times]).grid(row=2, rowspan=1, column=times, columnspan=1)
                            except:
                                pass  
                    #If filter for SKU                                                                                                                                                       
                    else:   
                        #Search by SKUID
                        if entryVal[0] != "":
                            cursor.execute("SELECT * FROM SKU WHERE SKUID=:skuid", {"skuid": entryVal[0]})
                        #Search by SKU Name
                        elif entryVal[1] != "":
                            cursor.execute("SELECT * FROM SKU WHERE SKUName LIKE :name", {"name": '%'+entryVal[1]+'%'})
                        
                        info = cursor.fetchall()
                        for row in range (len(info)):
                            for col in range (len(info[row])):
                                ttk.Label(content.frames[0], text=info[row][col]).grid(row=row+2, rowspan=1, column=col, columnspan=1)

            def view(page, ID):
                    if page == "GRN":
                            cursor.execute("SELECT POSKUID, ActArrDate FROM GRN WHERE GRNNo=:grnno", {"grnno": ID})
                            grnArr = cursor.fetchall()[0]
                            poskuid, actDate = grnArr[0], grnArr[1]
                            cursor.execute("SELECT POID, SKUID, Qty, TtlPurchasePrice, ExpArrDate FROM POSKU WHERE POSKUID=:poskuid", {"poskuid": poskuid})
                            appendArr = []
                            for row in cursor.fetchall():
                                temp = []
                                poid, skuid, qty, total, expDate = row[0], row[1], row[2], row[3], row[4]
                                cursor.execute("SELECT SupplierID, PODate FROM PurchaseOrder WHERE POID=:poid", {"poid": poid})
                                poArr = cursor.fetchall()[0]
                                suppID, podate = poArr[0], poArr[1]
                                cursor.execute("SELECT SupplierName FROM Supplier WHERE SupplierID=:suppID", {"suppID": suppID})
                                suppName = cursor.fetchall()[0][0]
                                cursor.execute("SELECT SKUName FROM SKU WHERE SKUID=:skuid", {"skuid": skuid})
                                temp.append(skuid)
                                temp.append(cursor.fetchall()[0][0])
                                temp.append(qty)
                                temp.append(total)
                                temp.append(expDate)
                                temp.append(actDate)
                                appendArr.append(temp)

                            widgetsArr = [[[ttk.Label(master, text="GRN No: "), 0, 1, 0, 1, 0], [ttk.Label(master, text=ID), 0, 1, 1, 1, 0], [ttk.Label(master, text="PO No: "), 0, 1, 2, 1, 0], [ttk.Label(master, text=poid), 0, 1, 3, 1, 0], [ttk.Label(master, text="Supplier: "), 0, 1, 4, 1, 0], [ttk.Label(master, text=suppName), 0, 1, 5, 1, 0], [ttk.Label(master, text="PO Date: "), 0, 1, 6, 1, 0], [ttk.Label(master, text=podate), 0, 1, 7, 1, 0], 
                                           [ttk.Label(master, text="SKUID"), 1, 1, 0, 1, 0], [ttk.Label(master, text="SKU Name"), 1, 1, 1, 1, 0], [ttk.Label(master, text="Qty"), 1, 1, 2, 1, 0], [ttk.Label(master, text="Total Purchase Price"), 1, 1, 3, 1, 0], [ttk.Label(master, text="Expected Arrival Date"), 1, 1, 4, 1, 0], [ttk.Label(master, text="Actual Arrival Date"), 1, 1, 5, 1, 0]]]

                    elif page == "SO":
                            cursor.execute("SELECT SODate FROM SalesOrder WHERE SOID=:soid", {"soid": ID})
                            sodate = cursor.fetchall()[0][0]
                            cursor.execute("SELECT * FROM SOSKU WHERE SOID=:soid", {"soid": ID})
                            total = 0
                            appendArr = []
                            for row in cursor.fetchall():
                                temp = []
                                soskuid, skuid, qty, adder, dvDate, dvStatus = row[0], row[2], row[3], row[4], row[5], row[6]
                                total += adder
                                cursor.execute("SELECT SKUName FROM SKU WHERE SKUID=:skuid", {"skuid": skuid})
                                skuname = cursor.fetchall()[0][0]
                                temp.append(skuid)
                                temp.append(skuname)
                                temp.append(qty)
                                temp.append(adder)
                                temp.append(dvDate)
                                temp.append(dvStatus)
                                appendArr.append(temp)
                                
                            widgetsArr = [[[ttk.Label(master, text="Order No: "), 0, 1, 0, 1, 0], [ttk.Label(master, text=ID), 0, 1, 1, 1, 0], [ttk.Label(master, text="SO Date:"), 0, 1, 2, 1, 0], [ttk.Label(master, text=sodate), 0, 1, 3, 1, 0], [ttk.Label(master, text="Total amount: "), 0, 1, 4, 1, 0], [ttk.Label(master, text=total), 0, 1, 5, 1, 0], 
                                           [ttk.Label(master, text="SKUID"), 1, 1, 0, 1, 0], [ttk.Label(master, text="SKU Name"), 1, 1, 1, 1, 0], [ttk.Label(master, text="Qty"), 1, 1, 2, 1, 0], [ttk.Label(master, text="Total Selling Price"), 1, 1, 3, 1, 0], [ttk.Label(master, text="Delivery Date"), 1, 1, 4, 1, 0], [ttk.Label(master, text="Delivery Status"), 1, 1, 5, 1, 0]]]

                    else:
                            cursor.execute("SELECT SupplierID, PODate, Approval FROM PurchaseOrder WHERE POID=:poid", {"poid": ID})
                            for row in cursor.fetchall():
                                suppID = row[0]
                                poDate = row[1]
                                approval = row[2]
                                if page == "PO":
                                    approval = row[2]
                                    if approval != "Rejected" or approval != "Pending":
                                        approval = "Approved"
                            cursor.execute("SELECT SupplierName FROM Supplier WHERE SupplierID=:supID", {"supID": suppID})
                            supplierName = cursor.fetchall()[0][0]
                            cursor.execute("SELECT SKUID, Qty, TtlPurchasePrice, ExpArrDate, ArrStatus FROM POSKU WHERE POID=:poid", {"poid": ID})
                            appendArr = []   #create an array to hold values to be appended into widgetsArr later
                            for row in cursor.fetchall():
                                temp = []
                                temp.append(row[0])
                                cursor.execute("SELECT SKUName FROM SKU WHERE SKUID=:skuid", {"skuid": row[0]})
                                skuName = cursor.fetchall()[0][0]
                                temp.append(skuName)
                                temp.append(row[1])
                                temp.append(row[2])
                                temp.append(row[3])
                                temp.append(row[4])
                                appendArr.append(temp)

                            if page == "PO":
                                widgetsArr = [[[ttk.Label(master, text="Order No: "), 0, 1, 0, 1, 0], [ttk.Label(master, text=ID), 0, 1, 1, 1, 0], [ttk.Label(master, text="Supplier"), 0, 1, 2, 1, 0], [ttk.Label(master, text=supplierName), 0, 1, 3, 1, 0], [ttk.Label(master, text="PO date"), 0, 1, 4, 1, 0], [ttk.Label(master, text=poDate), 0, 1, 5, 1, 0], [ttk.Label(master, text="Approval: "), 0, 1, 6, 1, 0], [ttk.Label(master, text=approval), 0, 1, 7, 1, 0], 
                                [ttk.Label(master, text="SKUID"), 1, 1, 0, 1, 0], [ttk.Label(master, text="SKU Name"), 1, 1, 1, 1, 0], [ttk.Label(master, text="QTY"), 1, 1, 2, 1, 0], [ttk.Label(master, text="Total Price"), 1, 1, 3, 1, 0], [ttk.Label(master, text="ETA"), 1, 1, 4, 1, 0], [ttk.Label(master, text="Arrival status"), 1, 1, 5, 1, 0]]]
                            else:
                                widgetsArr = [[[ttk.Label(master, text="Order No: "), 0, 1, 0, 1, 0], [ttk.Label(master, text=ID), 0, 1, 1, 1, 0], [ttk.Label(master, text="Supplier"), 0, 1, 2, 1, 0], [ttk.Label(master, text=supplierName), 0, 1, 3, 1, 0], [ttk.Label(master, text="PO date"), 0, 1, 4, 1, 0], [ttk.Label(master, text=poDate), 0, 1, 5, 1, 0], 
                                [ttk.Label(master, text="SKUID"), 1, 1, 0, 1, 0], [ttk.Label(master, text="SKU Name"), 1, 1, 1, 1, 0], [ttk.Label(master, text="QTY"), 1, 1, 2, 1, 0], [ttk.Label(master, text="Total Price"), 1, 1, 3, 1, 0], [ttk.Label(master, text="ETA"), 1, 1, 4, 1, 0], [ttk.Label(master, text="Arrival status"), 1, 1, 5, 1, 0]]]

                    for row in range (len(appendArr)):
                        for col in range (len(appendArr[row])):
                            widgetsArr[0].append([ttk.Label(master, text=appendArr[row][col]), row+2, 1, col, 1, 0])

                    def saveDb():
                        cursor.execute("UPDATE PurchaseOrder SET Done=1 WHERE POID=:poid;", {"poid": ID})
                        cursor.execute("SELECT Approval FROM PurchaseOrder WHERE POID=:poid;", {"poid": ID})
                        approval = cursor.fetchall()[0][0]
                        if approval != "Rejected":
                            cursor.execute("SELECT SKUID, Qty FROM POSKU WHERE POID=:poid", {"poid": poid})
                            for rows in cursor.fetchall():
                                cursor.execute("SELECT CurrentStockLvl FROM SKU WHERE SKUID=:skuid", {"skuid": rows[0]})
                                cursor.execute("UPDATE SKU SET CurrentStockLvl=:curr WHERE SKUID=:skuid", {"curr": rows[1]+cursor.fetchall()[0][0], "skuid": rows[0]})
                        connection.commit()
                        cls.__init__(master, self.width, self.height, cls.username, header=True)

                    if page == "Dashboard":
                        widgetsArr[0].append([ttk.Label(master, text="Approval: "), 0, 1, 6, 1, 0])
                        widgetsArr[0].append([ttk.Label(master, text=approval, foreground="red"), 0, 1, 7, 1, 0])
                        widgetsArr[0].append([ttk.Button(master, text="Done", command=saveDb), row+3, 1, 5, 1, 0])


                    view = cls.Content(cls, master, widgetsArr, cls.height*0.9, cls.width, 0, subframe=1)
                         
            match index:
                    #Change Layout.widgetsArr so the contentFrame can be updated according to page
                    case 0:   #Dashboard
                        cls.__init__(master, self.width, self.height, self.username)
                    case 1:   #Purchase
                        pageLbl.configure(text='Purchase Order')

                        startDate = tk.StringVar()
                        endDate = tk.StringVar()
                        start = ttk.Entry(master, textvariable=startDate)
                        end = ttk.Entry(master, textvariable=endDate)
                        cls.widgetsArr = [[[start, 0, 1, 0, 1, 0], [ttk.Label(master, text="To"), 0, 1, 1, 1, 5], [end, 0, 1, 2, 1, 0], [ttk.Label(master, text="Search PONo:"), 0, 1, 3, 1, 0], [ttk.Entry(master), 0, 1, 4, 1, 0], [ttk.Button(master, text="🔍", command=lambda: search("PO")), 0, 1, 5, 1, 0], [ttk.Button(master, text="+", command=lambda: addPurchase(cls, master)), 0, 1, 6, 1, 0],
                                           [ttk.Label(text="Order no."), 1, 1, 0, 1, 0], [ttk.Label(text="Vendor"), 1, 1, 1, 1, 0], [ttk.Label(text="Date Created"), 1, 1, 2, 1, 0], [ttk.Label(text="Approval status"), 1, 1, 3, 1, 0], [ttk.Label(text="Total price"), 1, 1, 4, 1, 0]]]
                        
                        #Fetch data from PO db and show on window
                        connection = self.__connectDb__(self.__database__)   #Connect to database
                        cursor = connection.cursor()
                        cursor.execute("SELECT POID, SupplierID, PODate, Approval FROM PurchaseOrder ORDER BY POID DESC LIMIT 50;")
                        po = cursor.fetchall()
                        for rowNo in range (len(po)):
                            row = po[rowNo]
                            for infoNo in range(len(row)):
                                if infoNo == 1:
                                    cursor.execute("SELECT SupplierName FROM Supplier WHERE SupplierID=:supID", {"supID": row[infoNo]})
                                    info = cursor.fetchall()[0][0]
                                elif infoNo == 3:
                                    if row[infoNo] != "Pending" or "Rejected":
                                        info = row[infoNo]
                                    else:
                                        info = "Approved by " + row[infoNo]
                                else:
                                    info = row[infoNo]
                                if infoNo == 0:
                                    temp = row[0]
                                    cls.widgetsArr[0].append([ttk.Button(master, text=info, command=lambda temp=temp: view("PO", temp)), rowNo+2, 1, infoNo, 1, 0])
                                    
                                else:
                                    cls.widgetsArr[0].append([ttk.Label(master, text=info), rowNo+2, 1, infoNo, 1, 0])

                            #Calculate total price of each PO
                            total = 0
                            cursor.execute("SELECT TtlPurchasePrice FROM POSKU WHERE POID=:poid", {"poid": row[0]})
                            for row in cursor.fetchall():
                                if row[0].__class__.__name__ == 'str':
                                    adder = row[0].strip()
                                    adder = adder.replace(',', '') 
                                    adder = float(adder)
                                else:
                                    adder = row[0]
                                total += adder
                            cls.widgetsArr[0].append([ttk.Label(master, text=str(round(total, 2))), rowNo+2, 1, infoNo+1, 1, 0])

                        purchase = cls.Content(cls, master, cls.widgetsArr, cls.height*0.9, cls.width, 0)
                    case 2:   #Receive
                        pageLbl.configure(text='Goods Receive')

                        startDate = tk.StringVar()
                        endDate = tk.StringVar()
                        start = ttk.Entry(master, textvariable=startDate)
                        end = ttk.Entry(master, textvariable=endDate)
                        cls.widgetsArr = [[[start, 0, 1, 0, 1, 0], [ttk.Label(master, text="To"), 0, 1, 1, 1, 5], [end, 0, 1, 2, 1, 0], [ttk.Label(master, text="Search GRNNo:"), 0, 1, 3, 1, 0], [ttk.Entry(master), 0, 1, 4, 1, 0], [ttk.Button(master, text="🔍", command=lambda: search("GRN")), 0, 1, 5, 1, 0], [ttk.Button(master, text="+", command=lambda: addReceive(cls, master)), 0, 1, 6, 1, 0],
                                           [ttk.Label(master, text="GRN No"), 1, 1, 0, 1, 0], [ttk.Label(master, text="SKUID"), 1, 1, 1, 1, 0], [ttk.Label(master, text="SKU Name"), 1, 1, 2, 1, 0], [ttk.Label(master, text="Qty"), 1, 1, 3, 1, 0], [ttk.Label(master, text="Actual Arrival Date"), 1, 1, 4, 1, 0]]]
                        
                        #Fetch GRN information
                        cursor.execute("SELECT * FROM GRN ORDER BY GRNNo DESC LIMIT 50;")
                        grnArr = cursor.fetchall()
                        rowindex=2
                        for row in grnArr:
                            grnNo, poskuID, actDate = row[0], row[1], row[2]
                            cursor.execute("SELECT POID, SKUID, Qty, TtlPurchasePrice FROM POSKU WHERE POSKUID=:poskuid", {"poskuid": poskuID})
                            poskuArr = cursor.fetchall()[0]
                            skuid, qty, total = poskuArr[1], poskuArr[2], poskuArr[3]
                            cursor.execute("SELECT SKUName FROM SKU WHERE SKUID=:skuid", {"skuid":skuid})
                            skuname = cursor.fetchall()[0][0]

                            #Show GRN information on window
                            cls.widgetsArr[0].append([ttk.Button(master, text=grnNo, command= lambda grnNo=grnNo: view("GRN", grnNo)), rowindex, 1, 0, 1, 0])
                            cls.widgetsArr[0].append([ttk.Label(master, text=skuid), rowindex, 1, 1, 1, 0])
                            cls.widgetsArr[0].append([ttk.Label(master, text=skuname), rowindex, 1, 2, 1, 0])
                            cls.widgetsArr[0].append([ttk.Label(master, text=qty), rowindex, 1, 3, 1, 0])
                            cls.widgetsArr[0].append([ttk.Label(master, text=actDate), rowindex, 1, 4, 1, 0])
                            rowindex += 1

                        receive = cls.Content(cls, master, cls.widgetsArr, cls.height*0.9, cls.width, 0)
                    case 3:   #Sales
                        pageLbl.configure(text='Sales Order')

                        startDate = tk.StringVar()
                        endDate = tk.StringVar()
                        start = ttk.Entry(master, textvariable=startDate)
                        end = ttk.Entry(master, textvariable=endDate)
                        cls.widgetsArr = [[[start, 0, 1, 0, 1, 0], [ttk.Label(master, text="To"), 0, 1, 1, 1, 5], [end, 0, 1, 2, 1, 0], [ttk.Label(master, text="Search SO number: "), 0, 1, 3, 1, 0], [ttk.Entry(master), 0, 1, 4, 1, 0], [ttk.Button(master, text="🔍", command=lambda: search("SO")), 0, 1, 5, 1, 0], [ttk.Button(master, text="+", command=lambda: addSales(cls, master)), 0, 1, 6, 1, 0],
                                           [ttk.Label(master, text="Order no."), 1, 1, 0, 1, 0], [ttk.Label(master, text="Date Created"), 1, 1, 1, 1, 0], [ttk.Label(master, text="Total Selling Price"), 1, 1, 2, 1, 0]]]
                        
                        #Fetch SO information
                        cursor.execute("SELECT SOID, SODate FROM SalesOrder ORDER BY SOID DESC LIMIT 50;")
                        soArr = cursor.fetchall()

                        #Calculate total selling price
                        rowindex = 2
                        for row in soArr:
                            soID = row[0]
                            soDate = row[1]
                            cursor.execute("SELECT TtlSellingPrice FROM SOSKU WHERE SOID=:soid;", {"soid":  soID})
                            total = cursor.fetchall()[0][0]
                            
                            #Show widget on window
                            cls.widgetsArr[0].append([ttk.Button(master, text=soID, command=lambda soID=soID:   view("SO", soID)), rowindex, 1, 0, 1, 0])
                            cls.widgetsArr[0].append([ttk.Label(master, text=soDate), rowindex, 1, 1, 1, 0])
                            cls.widgetsArr[0].append([ttk.Label(master, text=str(total)), rowindex, 1, 2, 1, 0])
                            
                            rowindex += 1     

                        sales = cls.Content(cls, master, cls.widgetsArr, cls.height*0.9, cls.width, 0)
                    case 4:   #Vendors
                        pageLbl.configure(text='Vendors')

                        cls.widgetsArr = [[[ttk.Button(master, text="+", command=lambda: addVendors(cls, master)), 0, 1, 4, 1, 0],
                                           [ttk.Label(master, text="Vendor ID"), 1, 1, 0, 1, 0], [ttk.Label(master, text="Vendor name"), 1, 1, 1, 1, 0]]]
                        
                        #FETCH VENDORS LIST FROM DB AND APPEND TO THE ARRAY
                        cursor.execute("SELECT * FROM Supplier ORDER BY SupplierID;")
                        rowindex = 2
                        for row in cursor.fetchall():
                            cls.widgetsArr[0].append([ttk.Label(master, text=row[0]), rowindex, 1, 0, 1, 0])
                            cls.widgetsArr[0].append([ttk.Label(master, text=row[1]), rowindex, 1, 1, 1, 0])
                            rowindex += 1

                        vendors = cls.Content(cls, master, cls.widgetsArr, cls.height*0.9, cls.width, 0)
                    case 5:   #Stock
                        pageLbl.configure(text='SKU')

                        cls.widgetsArr = [[[ttk.Label(master, text="SKUID:"), 0, 1, 0, 1, 0], [ttk.Entry(master), 0, 1, 1, 1, 0], [ttk.Label(master, text="SKUName:"), 0, 1, 2, 1, 0], [ttk.Entry(master), 0, 1, 3, 1, 0], [ttk.Button(master, text="🔍", command=lambda: search("SKU")), 0, 1, 4, 1, 0], [ttk.Button(master, text="+", command=lambda: addSKU(cls, master)), 0, 1, 5, 1, 0],
                                           [ttk.Label(master, text="SKUID"), 1, 1, 0, 1, 0], [ttk.Label(master, text="SKU Name"), 1, 1, 1, 1, 0], [ttk.Label(master, text="Current Stock Level"), 1, 1, 2, 1, 0], [ttk.Label(master, text="Opening Stock"), 1, 1, 3, 1, 0]]]
                        #Fetch SKU info
                        cursor.execute("SELECT * FROM SKU ORDER BY SKUID DESC LIMIT 50;")

                        #Show information on window
                        rowindex = 2
                        for row in cursor.fetchall():
                            cls.widgetsArr[0].append([ttk.Label(master, text=row[0]), rowindex, 1, 0, 1, 0])
                            cls.widgetsArr[0].append([ttk.Label(master, text=row[1]), rowindex, 1, 1, 1, 0])
                            cls.widgetsArr[0].append([ttk.Label(master, text=row[3]), rowindex, 1, 2, 1, 0])
                            cls.widgetsArr[0].append([ttk.Label(master, text=row[2]), rowindex, 1, 3, 1, 0])
                            rowindex += 1

                        stock = cls.Content(cls, master, cls.widgetsArr, cls.height*0.9, cls.width, 0)
                    case 6:   #Upload Discount Dates
                        def saveDb():
                            entryVal = []
                            for widget in date.frames[0].winfo_children():
                                if widget.winfo_class() == 'TEntry':
                                    entryVal.append(widget.get())   #Get the start and end dates for SD1 and SD2

                            #Start dates and end dates for different discount periods  
                            try:
                                summerStart = datetime.datetime.strptime(entryVal[0], r'%d/%m/%Y')
                                summerEnd = datetime.datetime.strptime(entryVal[1], r'%d/%m/%Y')
                                cnyStart = datetime.datetime.strptime(entryVal[2], r'%d/%m/%Y')
                                cnyEnd = datetime.datetime.strptime(entryVal[3], r'%d/%m/%Y')
                                easterStart = datetime.datetime.strptime(entryVal[4], r'%d/%m/%Y')
                                easterEnd = datetime.datetime.strptime(entryVal[5], r'%d/%m/%Y')
                                xmasStart = datetime.datetime.strptime(entryVal[6], r'%d/%m/%Y')
                                xmasEnd = datetime.datetime.strptime(entryVal[7], r'%d/%m/%Y')

                                adddate = datetime.datetime.strptime('01/01/'+str(summerStart.year), r'%d/%m/%Y')
                                while adddate.year == summerStart.year:
                                    vip = 0
                                    sd1 = 0
                                    sd2 = 0
                                    if (adddate>=cnyStart and adddate<=cnyEnd) or (adddate>=easterStart and adddate<=easterEnd) or (adddate>=xmasStart and adddate<=xmasEnd):   #If date is between cny or easter or xmas discount period
                                        sd1 = 1
                                    if adddate>=summerStart and adddate<=summerEnd:   #If date is between summer days
                                        sd2 = 1
                                    if adddate.weekday()+1 == 3 or adddate.weekday()+1 == 5:   #If date is wed or fri (vip day)
                                        vip = 1
                                    cursor.execute("INSERT INTO SPECIAL VALUES(?, ?, ?, ?)", (adddate.strftime(r'%d/%m/%Y'), vip, sd1, sd2))
                                    adddate += datetime.timedelta(days=1)
                                
                                connection.commit()
                            except:
                                ttk.Label(date.frames[0], text="Invalid Entry", foreground="red").grid(row=8, column=0)

                        pageLbl.configure(text="Upload Date")

                        #Add widget labels to window
                        cls.widgetsArr = [[[ttk.Label(master, text="Summer Days Start Date: "), 0, 1, 0, 1, 0],
                                           [ttk.Entry(master), 0, 1, 1, 1, 0],
                                           [ttk.Label(master, text="Summer Days End Date: "), 1, 1, 0, 1, 0],
                                           [ttk.Entry(master), 1, 1, 1, 1, 0], 
                                           [ttk.Label(master, text="CNY Discount Start Date: "), 2, 1, 0, 1, 0],
                                           [ttk.Entry(master), 2, 1, 1, 1, 0], 
                                           [ttk.Label(master, text="CNY Discount End Date: "), 3, 1, 0, 1, 0],
                                           [ttk.Entry(master), 3, 1, 1, 1, 0], 
                                           [ttk.Label(master, text="Easter Discount Start Date: "), 4, 1, 0, 1, 0],
                                           [ttk.Entry(master), 4, 1, 1, 1, 0], 
                                           [ttk.Label(master, text="Easter Discount End Date: "), 5, 1, 0, 1, 0],
                                           [ttk.Entry(master), 5, 1, 1, 1, 0], 
                                           [ttk.Label(master, text="Christmas Discount Start Date: "), 6, 1, 0, 1, 0],
                                           [ttk.Entry(master), 6, 1, 1, 1, 0], 
                                           [ttk.Label(master, text="Christmas Discount End Date: "), 7, 1, 0, 1, 0],
                                           [ttk.Entry(master), 7, 1, 1, 1, 0], 
                                           [ttk.Button(master, text="Save", command=saveDb), 8, 1, 1, 1, 0]]]
                        
                        date = cls.Content(cls, master, cls.widgetsArr, cls.height*0.6, cls.width, 0)

        #Buttons for all pages
        dashboardBtn = ttk.Button(menuFrame, text="Dashboard", command=lambda: btnClick(self, master, root, 0))
        purchaseBtn = ttk.Button(menuFrame, text="Purchase Order", command=lambda: btnClick(self, master, root, 1))
        receiveBtn = ttk.Button(menuFrame, text="Goods Receive", command=lambda: btnClick(self, master, root, 2))
        salesBtn = ttk.Button(menuFrame, text="Sales Order", command=lambda: btnClick(self, master, root, 3))
        vendorsBtn = ttk.Button(menuFrame, text="Vendors", command=lambda: btnClick(self, master, root, 4))
        skuBtn = ttk.Button(menuFrame, text="SKU", command=lambda: btnClick(self, master, root, 5))
        dateBtn = ttk.Button(menuFrame, text="Upload date", command=lambda: btnClick(self, master, root, 6))
        dashboardBtn.grid(row=0, column=1, sticky='e')
        purchaseBtn.grid(row=0, column=2, sticky='e')
        receiveBtn.grid(row=0, column=3, sticky='e')
        salesBtn.grid(row=0, column=4, sticky='e')
        vendorsBtn.grid(row=0, column=5, sticky='e')
        skuBtn.grid(row=0, column=6, sticky='e')
        dateBtn.grid(row=0, column=7, sticky='e')

    class Content(ttk.Frame):
        def __init__(self, layout, master, widgets, height, width, variable, subframe=1):
            self.master = master
            self.subframe = subframe
            self.contentFrame = ttk.Frame(master)
            self.layout = layout
            try:
                #Destroy original contents
                layout.frame[2].destroy()
                layout.frame.pop(2)
            except:
                pass
            layout.frame.append(self.contentFrame)
            self.contentFrame.grid(row=2)
            self.contentFrame.grid_propagate(0)
            self.frames = []
            self.subframes(self.subframe, widgets, height, width, variable)

        #Add different columns
        def subframes(self, subframeNo, widgets, height, width, variable):   #widgets is a 3d array of the widgets present in the subframe
            #Clone the widget into the subframe
            def clone(master, widget, variable):   
                cls = widget.__class__
                if cls == tk.ttk.OptionMenu:
                    clone = cls(master, variable)
                else:
                    clone = cls(master)
                for key in widget.configure():
                    try:
                        if key == "compound":
                            clone.configure({"compound": "none"})
                        else:
                            clone.configure({key: widget.cget(key)})
                    except:   #There are some read-only keys which will be passed when creating the widget
                        continue
                return clone
            
            for j in range (subframeNo):
                s = ttk.Style(self.master)
                s.configure('Frame1.TFrame', background='light blue')
                s.configure('Frame2.TFrame', background='MediumPurple1')
                subFrame = ttk.Frame(self.contentFrame, height=height*(1-0.1), width=width/subframeNo)
                subFrame.grid_propagate(0)
                subFrame.pack(side="left")
                wdgtarr = widgets[j]
                
                #Create a scrollbar
                def sty(index):
                    if index%2 == 0:
                        return("Frame1.TFrame")
                    else:
                        return("Frame2.TFrame")
                #Create canvas to scroll
                scbCanvas = tk.Canvas(subFrame, width=(width-100)/subframeNo, height=height*0.9)
                scbCanvas.pack(side="left", fill='both', expand=1)
                scb = ttk.Scrollbar(subFrame, orient="vertical")
                self.yScrollBar = scb
                self.canvas = scbCanvas
                scb.pack(side="right", fill="y")
                scb.config(command=scbCanvas.yview)
                self.canvas.configure(yscrollcommand=scb.set)

                onSubFrame = ttk.Frame(self.canvas, style=sty(j))   #Frame on canvas including all contents
                self.canvasFrame = onSubFrame
                scbCanvas.create_window((0,0), window=onSubFrame, anchor="nw")

                for i in range(len(wdgtarr)):
                    wdgtData = wdgtarr[i]
                    wdgt = clone(onSubFrame, wdgtData[0], variable)
                    wdgt.grid(row=wdgtData[1], rowspan=wdgtData[2], column=wdgtData[3], columnspan=wdgtData[4], ipadx=wdgtData[5])

                self.canvas.bind('<Configure>', lambda e:scbCanvas.configure(scrollregion = scbCanvas.bbox("all")))   #Bind canvas to scrollbar
                self.frames.append(onSubFrame)


                


#-------Main program-------#
root = tk.Tk()
root.title("Stock Property Management System")
root.geometry("600x600")
root.minsize(600,600)
login = Login(root)
root.mainloop()